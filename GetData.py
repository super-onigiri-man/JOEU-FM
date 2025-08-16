import requests
from bs4 import BeautifulSoup #スクレイピング（取得）用
import datetime #日付計算用
import openpyxl
import sqlite3
import PySimpleGUI as sg
import asyncio #非同期処理
import unicodedata #全角文字を半角文字に変換
import re
import xlrd #NewHaruyaPath用


dbname = ('test.db')
conn = sqlite3.connect(dbname, isolation_level=None)#データベースを作成、自動コミット機能ON
cursor = conn.cursor() #カーソルオブジェクトを作成

# テーマカラーを設定
sg.theme('SystemDefault')
#scoreを元にしたオリコンシングル・オリコンデジタル・ビルボード総合ランキング

# オリコン週間ランキング用
OriconWeekData = []
# オリコンデジタルランキング用
OriconDigitalData = []
# ビルボードランキング用
BillboardData = []
# 明屋書店ランキング用
HaruyaData = []

def clear():
    OriconWeekData.clear()
    OriconDigitalData.clear()
    BillboardData.clear()
    HaruyaData.clear()

# 初めて処理を行うかどうかのフラグ
popup_done = False

def OriconTodays():
    global popup_done
    
    # 今日の日付と曜日を求める
    dt = datetime.date.today()
    weekday = dt.weekday()

    if weekday <= 1:  # 月曜日または火曜日（今週月曜日の日付を返す）
        Oriconday = dt - datetime.timedelta(days=(weekday))
    elif weekday == 2:  # 水曜日（時間を判定する）
        current_time = datetime.datetime.now().time()
        specified_time = datetime.time(14, 10)  
        if current_time < specified_time and not popup_done:
            sg.popup('先週のランキングを取得します\n今週のランキングは本日14:10以降に取得可能です',no_titlebar=True)
            popup_done = True  # ポップアップが表示されたことをフラグで管理
            Oriconday = dt - datetime.timedelta(days=2) # 14:10までは先週のデータを取得
        elif current_time < specified_time:
            Oriconday = dt - datetime.timedelta(days=2) # 14:10までは先週のデータを取得
        else:
            Oriconday = dt + datetime.timedelta(days=5) # 14:10以降は今週のデータを取得
    else:  # 木曜日から日曜日（来週月曜日の日付を返す）
        Oriconday = dt + datetime.timedelta(days=(7 - weekday))

    WriteLog(2,str(dt)+"から"+str(Oriconday)+"に変換")

    return Oriconday


def Flags():

    # 今日の日付と曜日を求める
    dt = datetime.date.today()
    weekday = dt.weekday()

    if weekday <= 1:  # 月曜日または火曜日（今週月曜日の日付を返す）
        Flag = True
    elif weekday == 2:  # 水曜日（時間を判定する）
        current_time = datetime.datetime.now().time()
        specified_time = datetime.time(14, 10)  
        if current_time < specified_time and not popup_done:
            Flag = False
        elif current_time < specified_time:
            Flag = False
        else:
            Flag = True # 14:10以降は今週のデータを取得
    else:  # 木曜日から日曜日（来週月曜日の日付を返す）
        Flag = True

    return Flag

def OriconLastWeek():
    # 今日の日付と曜日を求める
    dt = datetime.date.today()
    weekday = dt.weekday()
    if (weekday == 0):  # 今日が月曜日(先週(今週月曜日)のランキング表示)
        Oriconday = dt - datetime.timedelta(days=7)
    elif (weekday == 1):  # 今日が火曜日(先週(今週月曜日)のランキング表示)
        Oriconday = dt - datetime.timedelta(days=8)
    elif (weekday >= 2):  # 今日が水曜日以降
        Oriconday = dt - datetime.timedelta(days=(weekday +7) % 7)

    WriteLog(2,str(dt)+"から"+str(Oriconday)+"に変換")

    return Oriconday




def OriconSelectWeek(SelectDay):

    # 入力された日付をdatetimeオブジェクトに変換
    dt = SelectDay
    dt = dt.date()
    # 曜日を取得
    weekday = dt.weekday()
    if (weekday == 0):  # 今日が月曜日(先週(今週月曜日)のランキング表示)
        Oriconday = dt - datetime.timedelta(days=7)
    elif (weekday == 1):  # 今日が火曜日(先週(今週月曜日)のランキング表示)
        Oriconday = dt - datetime.timedelta(days=8)
    elif (weekday >= 2):  # 今日が水曜日以降
        Oriconday = dt - datetime.timedelta(days=(weekday +7) % 7)

    WriteLog(2,str(dt)+"から"+str(Oriconday)+"に変換")

    return Oriconday

# 独自ID自動生成
def generate_unique_id(song_title, artist_name):
    # 曲名とアーティスト名の頭3文字を取得
    song_prefix = song_title[:3]
    artist_prefix = artist_name[:3]

    # 頭3文字をUnicodeコードポイントに変換
    song_code_points = [ord(char) for char in song_prefix]
    artist_code_points = [ord(char) for char in artist_prefix]

    # コードポイントを文字列に変換して結合
    song_code_str = ''.join(f'{cp:04x}' for cp in song_code_points)
    artist_code_str = ''.join(f'{cp:04x}' for cp in artist_code_points)

    # 独自IDを生成
    unique_id = song_code_str + artist_code_str

    return unique_id



def OriconWeekRank(Oriconday):#オリコン週間ランキング

    try:
      
        while True:
            # print(count)
            # print(Oriconday)
            load_url = "https://www.oricon.co.jp/rank/js/w/" + str(Oriconday) + "/"
            html = requests.get(load_url)
            soup = BeautifulSoup(html.text, 'lxml')
            main_content = soup.find(class_="content-rank-main")
            if main_content:
                break

            # 日付を一日減らす
            Oriconday = Oriconday - datetime.timedelta(days=1)

            # count = count + 1

        #1位から10位
        links = soup.find(class_="content-rank-main").find_all('h2',class_='title') #曲名
        artist = soup.find(class_="content-rank-main").find_all('p',class_='name') #アーティスト名
        score = 6.0 #独自スコア
        rank = 1 #ランキング

        index = 0
        for link, artist in zip(links, artist):
            if '/' in link.text:
                # If yes, split the title and create two separate entries in the array
                titles = link.text.split('/')
                for title in titles:
                    unicodedata.normalize("NFKC", title.strip()) 
                    unicodedata.normalize("NFKC", artist.text)  
                    OriconWeekData.append([title.strip(), artist.text, format(score, '.1f'),generate_unique_id(title.strip(),artist.text)])

                    if title == titles[-1]:
                        rank = rank + 1
                        score = score - 0.3
            else:
                # If no slash, just add the entry to the array
                unicodedata.normalize("NFKC", link.text)  # Strip to remove leading/trailing whitespaces
                unicodedata.normalize("NFKC", artist.text)
                OriconWeekData.append([link.text, artist.text, format(score, '.1f'),generate_unique_id(link.text,artist.text)])
                rank = rank + 1
                score = score - 0.3

        # 壊れたときの表示用
        # if rank != 10:#10位以下（１ケタの場合）なら（点数の位置を揃えるため）
        #   print(" " + str(rank) + "位 " + "{:.1f}　 ".format(score) + link.text + "/" + artist.text)
        # else: #10位（２ケタの場合）なら
        #   print(str(rank) + "位 " + "{:.1f}　 ".format(score) + link.text + "/" + artist.text)

        # print(OriconWeekData)

        #11位から20位
        load_url = "https://www.oricon.co.jp/rank/js/w/" + str(Oriconday) + "/p/2/"
        html = requests.get(load_url)
        soup = BeautifulSoup(html.text, 'lxml')
        links = soup.find(class_="content-rank-main").find_all('h2', class_='title')  # 曲名
        artist = soup.find(class_="content-rank-main").find_all('p', class_='name')  # アーティスト名
        for link, artist in zip(links, artist):
            if '/' in link.text:
            # 「楽曲A/楽曲B」のような表現方法の場合
                titles = link.text.split('/') #"/"で２曲に分ける
                for title in titles:
                    unicodedata.normalize("NFKC", title.strip()) 
                    unicodedata.normalize("NFKC", artist.text)
                    OriconWeekData.append([title.strip(), artist.text, format(score, '.1f'),generate_unique_id(title.strip(),artist.text)])

                    if title == titles[-1]:
                        rank = rank + 1
                        score = score - 0.3
            else:
                # If no slash, just add the entry to the array
                unicodedata.normalize("NFKC", link.text)  # Strip to remove leading/trailing whitespaces
                unicodedata.normalize("NFKC", artist.text)
                OriconWeekData.append([link.text, artist.text, format(score, '.1f'),generate_unique_id(link.text,artist.text)])
                rank = rank + 1
                score = score - 0.3
                # 壊れたときの表示用
            # print(str(rank) + "位 " + "{:.1f}　 ".format(score) + link.text + "/" + artist.text)

        # print(OriconWeekData)
        print(str(Oriconday) + "付けオリコン週間シングルランキングOK")
        WriteLog(2,str(Oriconday)+"付けオリコン週間シングルランキング取得")


    except Exception as e:
        import traceback
        with open('error.log', 'a') as f:
            traceback.print_exc( file=f)
        WriteLog(4,'「オリコン週間シングルランキング」取得でエラー')
        sg.popup_error("「オリコン週間ランキング」が取得できませんでした",title="エラー",no_titlebar=True)


def OriconDigitalRank(Oriconday):#オリコンデジタルシングルランキング

    try:

        while True:
            # print(count)
            # print(Oriconday)
            load_url = "https://www.oricon.co.jp/rank/dis/w/" + str(Oriconday) + "/"
            html = requests.get(load_url)
            soup = BeautifulSoup(html.text, 'lxml')
            main_content = soup.find(class_="content-rank-main")
            if main_content:
                break

            # 日付を一日減らす
            Oriconday = Oriconday - datetime.timedelta(days=1)

            # count = count + 1

        # 1位から10位
        links = soup.find(class_="content-rank-main").find_all('h2', class_='title')
        artist = soup.find(class_="content-rank-main").find_all('p', class_='name')  # アーティスト名
        rank = 1
        score = float(6.0)
        for link, artist in zip(links, artist):
        # 壊れたときの表示用
            # if rank != 10:#10位以下（１ケタの場合）なら（点数の位置を揃えるため）
            #     print(" " + str(rank) + "位 " + "{:.1f}　 ".format(score) + link.text + "/" + artist.text)
            # else: #10位（２ケタの場合）なら
            #     print(str(rank) + "位 " + "{:.1f}　 ".format(score) + link.text + "/" + artist.text)
            unicodedata.normalize("NFKC", link.text)  # Strip to remove leading/trailing whitespaces
            unicodedata.normalize("NFKC", artist.text)
            OriconDigitalData.append([link.text,artist.text,format(score, '.1f'),generate_unique_id(link.text,artist.text)])
            rank = rank + 1
            score = score - 0.3

        # 11位から20位
        load_url = "https://www.oricon.co.jp/rank/dis/w/" + str(Oriconday) + "/p/2/"
        html = requests.get(load_url)
        soup = BeautifulSoup(html.text, 'lxml')
        links = soup.find(class_="content-rank-main").find_all('h2', class_='title')
        artist = soup.find(class_="content-rank-main").find_all('p', class_='name')  # アーティスト名
        for link, artist in zip(links, artist):
            unicodedata.normalize("NFKC", str(links))  # Strip to remove leading/trailing whitespaces
            unicodedata.normalize("NFKC", str(artist))
            OriconDigitalData.append([link.text,artist.text,format(score, '.1f'),generate_unique_id(link.text,artist.text)])
            # print(str(rank) + "位 " + "{:.1f}　 ".format(score) + link.text + "/" + artist.text)
            rank = rank + 1
            score = score - 0.3

        # print(OriconDigitalData)

        print(str(Oriconday) + "付けオリコンデジタルランキングOK")
        WriteLog(2,str(Oriconday)+"付けオリコンデジタルランキング取得")

    except Exception as e:
        import traceback
        with open('error.log', 'a') as f:
            traceback.print_exc( file=f)
        WriteLog(4,'「オリコンデジタルランキング取得」でエラー')
        sg.popup_error("「オリコンデジタルランキング」が取得できませんでした",title="エラー",no_titlebar=True)


def BillboadRank(Oriconday):#ビルボードJAPAN HOT100ランキング

    try:

        
        # URLの日付はOricondayと同じ

        Emergency = False

        while True:
            
            #URL(ここを変更すると読み込まなくなります)
            url = 'https://www.billboard-japan.com/charts/detail?a=hot100&year='+str(Oriconday.year)+'&month='+str(Oriconday.month)+'&day='+str(Oriconday.day)
            #URLを取得してくる
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            songs = soup.find_all('p', class_='musuc_title') #曲名
            if songs:
                break
                # count = count + 1

            Oriconday = Oriconday - datetime.timedelta(days=1)
            Emergency = True

        songs = soup.find_all('p', class_='musuc_title') #曲名
        artists = soup.find_all('p', class_='artist_name') #アーティスト名
        score = 6.0 #基準点

        for i in range(20):#20回繰り替えす
            song = str(songs[i].text.strip())
            artist = str(artists[i].text.strip())
            unicodedata.normalize("NFKC", song)  # Strip to remove leading/trailing whitespaces
            unicodedata.normalize("NFKC", artist)
            BillboardData.append([song,artist,format(score, '.1f'),generate_unique_id(song,artist)])
            # if i < 9:
            #   print(f" {i+1}位: {format(score, '.1f')} {song} / {artist}") #1位から9位までのランキング
            # else:
            #   print(f"{i+1}位: {format(score, '.1f')} {song} / {artist}") #10位から20位までのランキング
            score = score - 0.3 #scoreを-0.3する

        if Emergency:
            sg.popup('今週のランキングがなかったため最新日のランキングを取得しました',no_titlebar=True)

        # オリコンの日付とビルボードの発表日の差を埋めるための計算(表示用)
        Billday = Oriconday - datetime.timedelta(days=5)
        print(str(Billday) + "付けビルボードJAPAN HOT100ランキングOK")
        WriteLog(2,str(Billday)+"付けビルボードJAPAN HOT100ランキング取得")

        # print(BillboardData)

    except Exception as e:
        import traceback
        with open('error.log', 'a') as f:
            traceback.print_exc( file=f)

        WriteLog(4,'「ビルボードランキング」取得でエラー')
        sg.popup_error("「ビルボードランキング」が取得できませんでした",title="エラー",no_titlebar=True)


async def OldHaruyaRank(HaruyaPath): # 2024年6月末までの明屋書店フォーマット
    try:

        # Excelファイルを読み込む
        wb = openpyxl.load_workbook(HaruyaPath)
        ws = wb.active

        # データを2次元配列に挿入する
        for row in range(4, 24):
            song_names = unicodedata.normalize("NFKC", str(ws[f"D{row}"].value)).split('/')
            artist_name = unicodedata.normalize("NFKC", str(ws[f"C{row}"].value))
            # song_name = str(song_name)
            # artist_name = str(artist_name)
            point = round(6.0 - ((row - 4) * 0.3), 2)  # 点数を計算する
            for song_name in song_names:
                if "/" in song_name:
                    song_name_a, song_name_b = song_name.split("/")
                    str(song_name_a)
                    str(song_name_b)
                    HaruyaData.append([song_name_a.strip(), artist_name, point,generate_unique_id(song_name_a.strip(),artist_name)])
                    HaruyaData.append([song_name_b.strip(), artist_name, point,generate_unique_id(song_name_b.strip(),artist_name)])
                else:
                    HaruyaData.append([song_name.strip(), artist_name, point,generate_unique_id(song_name.strip(),artist_name)])

        print('旧・明屋書店データOK')
        WriteLog(2,"旧・明屋書店データ取得")

    except Exception as e:
            import traceback
            with open('error.log', 'a') as f:
                traceback.print_exc(file=f)
            WriteLog(4,'「旧・明屋書店」ランキング取得でエラー')
            sg.popup_error("「明屋書店ランキング」が取得できませんでした", title="エラー",no_titlebar=True)

async def NewHaruyaRank(HaruyaPath): # 2024年7月からの明屋書店フォーマット
    
    try:
        # Excelファイルの読み込み
        workbook = xlrd.open_workbook(HaruyaPath)
        sheet = workbook['SG1～20']

        # 既に登録された曲名を格納するセット
        registered_songs = set()
        point = 6.0

        # データの取得と整形
        for row_index in range(sheet.nrows):

            if point <= 0.3:
                break

            title = sheet.cell_value(row_index+5, 4)  # 曲名
            artist = sheet.cell_value(row_index+5, 5)  # アーティスト名

            # 全角・半角変換
            title = unicodedata.normalize("NFKC", title)
            artist = unicodedata.normalize("NFKC", artist)

            # print(title)
            # 正規表現で不要な文字列を削除
            title = re.sub(r'\(.*?\)', '', title)  # カッコとその中身を削除 
            title = re.sub(r'【.*?】', '',title) #デカカッコとその中身を削除
            title = re.sub(r'<.*>','',title)# <>を消す
            title = re.sub(r'[★▲]', '', title)  # 先頭の★と▲を削除

            # print(title)

            if "/" in title:
                title_a,title_b = title.split("/")
                str(title_a)
                str(title_b)
                if title_a not in registered_songs and title_b not in registered_songs:
                    # なかったらAを追加
                    # 独自ID設定
                    Unique_id_a = generate_unique_id(title_a,artist)
                    HaruyaData.append([title_a,artist,format(point, '.1f'),Unique_id_a])
                    registered_songs.add(title_a)

                    Unique_id_b = generate_unique_id(title_b,artist)
                    HaruyaData.append([title_b,artist,format(point, '.1f'),Unique_id_b])
                    registered_songs.add(title_b)
                    point = point - 0.3

            else:      
                # 同じ曲が登録済みかどうかを確認
                if title not in registered_songs:
                    # 新しい曲の場合、配列に追加
                    # 独自ID設定
                    Unique_id = generate_unique_id(title,artist)
                    HaruyaData.append([title,artist,format(point, '.1f'),Unique_id])
                    registered_songs.add(title)
                    point = point - 0.3

        # print(HaruyaData)
        print('新・明屋書店データOK')
        WriteLog(2,"新・明屋書店データ取得")


    except Exception as e:
            import traceback
            with open('error.log', 'a') as f:
                traceback.print_exc(file=f)
            WriteLog(4,'「新・明屋書店」ランキング取得でエラー')
            sg.popup_error("「明屋書店ランキング」が取得できませんでした", title="エラー",no_titlebar=True)

async def insertOriconWeekData():
    try:
        for entry in OriconWeekData:
            title, artist, score, unique_id= entry
            # 既存のデータがあればScoreを足して更新、なければ新規追加
            cursor.execute('''
                INSERT INTO music_master (Title, Artist, Score, Last_Rank, Last_Number, On_Chart ,Unique_id)
                VALUES (?, ?, ?, 0, 0, 0, ?)
                ON CONFLICT(Unique_id) DO UPDATE SET Score = Score + ?
            ''',(title,artist, score, unique_id,score))

            conn.commit()

        WriteLog(2,"オリコン週間シングルランキングをデータベースへ挿入")
    except Exception as e:
        import traceback
        with open('error.log', 'a') as f:
            traceback.print_exc( file=f)
        WriteLog(4,'オリコン週間シングルランキングをデータベースへ挿入でエラー')
        sg.popup_error("データベースを更新できませんでした。",title="エラー",no_titlebar=True)

async def insertOriconDigitalData():
    try:
        for entry in OriconDigitalData:
            title, artist, score, unique_id= entry
            # 既存のデータがあればScoreを足して更新、なければ新規追加
            cursor.execute('''
                INSERT INTO music_master (Title, Artist, Score, Last_Rank, Last_Number, On_Chart ,Unique_id)
                VALUES (?, ?, ?, 0, 0, 0, ?)
                ON CONFLICT(Unique_id) DO UPDATE SET Score = Score + ?
            ''', (title,artist, score, unique_id,score))

            conn.commit()

        WriteLog(2,"オリコンデジタルランキングをデータベースへ挿入")

    except Exception as e:
        import traceback
        with open('error.log', 'a') as f:
            traceback.print_exc( file=f)
            WriteLog(4,'オリコンデジタルランキングをデータベースへ挿入でエラー')
        sg.popup_error("データベースを更新できませんでした。",title="エラー",no_titlebar=True)

async def insertBillboardData():
    try:
        for entry in BillboardData:
            title, artist, score, unique_id= entry
            # 既存のデータがあればScoreを足して更新、なければ新規追加
            cursor.execute('''
                INSERT INTO music_master (Title, Artist, Score, Last_Rank, Last_Number, On_Chart ,Unique_id)
                VALUES (?, ?, ?, 0, 0, 0, ?)
                ON CONFLICT(Unique_id) DO UPDATE SET Score = Score + ?
            ''', (title,artist, score, unique_id,score))

            conn.commit()

        WriteLog(2,"ビルボードJAPAN HOT100をデータベースへ挿入")
    except Exception as e:
        import traceback
        with open('error.log', 'a') as f:
            traceback.print_exc( file=f)

        WriteLog(4,'ビルボードJAPAN HOT100をデータベースへ挿入でエラー')
        sg.popup_error("データベースを更新できませんでした。",title="エラー",no_titlebar=True)

async def insertHaruyaData():
    try:
        for entry in HaruyaData:
            title, artist, score, unique_id= entry
            # 既存のデータがあればScoreを足して更新、なければ新規追加
            cursor.execute('''
                INSERT INTO music_master (Title, Artist, Score, Last_Rank, Last_Number, On_Chart ,Unique_id)
                VALUES (?, ?, ?, 0, 0, 0, ?)
                ON CONFLICT(Unique_id) DO UPDATE SET Score = Score + ?
            ''', (title,artist, score, unique_id,score))

            conn.commit()

        WriteLog(2,"明屋書店データをデータベースへ挿入")

    except Exception as e:
        import traceback
        with open('error.log', 'a') as f:
            traceback.print_exc( file=f)
        WriteLog(4,'明屋書店データをデータベースへ挿入でエラー')
        sg.popup_error("データベースを更新できませんでした。",title="エラー",no_titlebar=True)

def GetThisWeekRank(HaruyaPath,Flag):

    WriteLog(1,"今週のデータ（明屋書店データあり）を選択")

    # レイアウト（共通設定）
    layout = [
        [sg.Text('読み込み中...', size=(15, 1)), sg.ProgressBar(72, orientation='h', size=(20, 20), key='progressbar')],
        [sg.Button('読み込み中止'),sg.Text(key = 'progmsg')]
    ]


    window = sg.Window('今週のランキング取得', layout,finalize=True,icon='FM-BACS.ico')

    def update_progress_bar(progress_bar,progmsg, value,msg):
        progress_bar.update_bar(value)
        progmsg.update(msg)
        window.refresh()

    while True:
        update_progress_bar(window['progressbar'],window['progmsg'],0,'データリセット中')
        clear()
        update_progress_bar(window['progressbar'],window['progmsg'], 0,'リロード用情報更新中')
        with open('Reload.txt', 'w',encoding='UTF-8') as f:
            f.write('1,'+HaruyaPath+',,'+str(Flag))
            # print(Flag)
        update_progress_bar(window['progressbar'],window['progmsg'], 8,'日付取得中')
        Oriconday=OriconTodays()
        update_progress_bar(window['progressbar'],window['progmsg'], 16,str(Oriconday)+'付けオリコン週間ランキング取得中')
        OriconWeekRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 24,str(Oriconday)+'付けオリコンデジタルランキング取得中')
        OriconDigitalRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 32,str(Oriconday - datetime.timedelta(days=5))+'付けビルボードランキング取得中')
        BillboadRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 40,'明屋書店ランキング取得中')
        if str(Flag) == 'True':
            asyncio.run(NewHaruyaRank(HaruyaPath))
        else:
            asyncio.run(OldHaruyaRank(HaruyaPath))
        update_progress_bar(window['progressbar'],window['progmsg'], 48,'DB登録・集計中')
        asyncio.run(insertOriconWeekData())    
        asyncio.run(insertOriconDigitalData())
        asyncio.run(insertBillboardData())
        asyncio.run(insertHaruyaData())
        
        window.close()

        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'キャンセル':
                WriteLog(1,"得点取得画面を閉じる操作")
                break

    window.close()

def NGetThisWeekRank():

    WriteLog(1,"今週のデータ（明屋書店データなし）を選択")

    # レイアウト（共通設定）
    layout = [
        [sg.Text('読み込み中...', size=(15, 1)), sg.ProgressBar(72, orientation='h', size=(20, 20), key='progressbar')],
        [sg.Button('読み込み中止'),sg.Text(key = 'progmsg')]
    ]


    window = sg.Window('今週のランキング取得', layout,finalize=True,icon='FM-BACS.ico')

    def update_progress_bar(progress_bar,progmsg, value,msg):
        progress_bar.update_bar(value)
        progmsg.update(msg)
        window.refresh()

    while True:
        update_progress_bar(window['progressbar'],window['progmsg'],0,'データリセット中')
        clear()
        update_progress_bar(window['progressbar'],window['progmsg'], 0,'リロード用情報更新中')
        with open('Reload.txt', 'w',encoding='UTF-8') as f:
            f.write('1,No,,')
            # print(Flag)
        update_progress_bar(window['progressbar'],window['progmsg'], 8,'日付取得中')
        Oriconday=OriconTodays()
        update_progress_bar(window['progressbar'],window['progmsg'], 16,str(Oriconday)+'付けオリコン週間ランキング取得中')
        OriconWeekRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 24,str(Oriconday)+'付けオリコンデジタルランキング取得中')
        OriconDigitalRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 32,str(Oriconday - datetime.timedelta(days=5))+'付けビルボードランキング取得中')
        BillboadRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 48,'DB登録・集計中')
        asyncio.run(insertOriconWeekData())    
        asyncio.run(insertOriconDigitalData())
        asyncio.run(insertBillboardData())
        
        window.close()

        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'キャンセル':
                WriteLog(1,"得点取得画面を閉じる操作")
                break

    window.close()


def GetLastWeekRank():

    WriteLog(1,"先週のデータ取得を選択")

    # レイアウト（共通設定）
    layout = [
        [sg.Text('読み込み中...', size=(15, 1)), sg.ProgressBar(72, orientation='h', size=(20, 20), key='progressbar')],
        [sg.Button('読み込み中止'),sg.Text(key = 'progmsg')]
    ]


    window = sg.Window('先週のランキング取得', layout,finalize=True,icon='FM-BACS.ico')

    def update_progress_bar(progress_bar,progmsg, value,msg):
        progress_bar.update_bar(value)
        progmsg.update(msg)
        window.refresh()

    while True:
        update_progress_bar(window['progressbar'],window['progmsg'],0,'データリセット中')
        clear()
        update_progress_bar(window['progressbar'],window['progmsg'], 0,'リロード用情報更新中')
        with open('Reload.txt', 'w',encoding='UTF-8') as f:
            f.write('2,,,')
        update_progress_bar(window['progressbar'],window['progmsg'], 8,'日付取得中')
        Oriconday=OriconLastWeek()
        update_progress_bar(window['progressbar'],window['progmsg'], 16,str(Oriconday)+'付けオリコン週間ランキング取得中')
        OriconWeekRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 24,str(Oriconday)+'付けオリコンデジタルランキング取得中')
        OriconDigitalRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 32,str(Oriconday - datetime.timedelta(days=5))+'付けビルボードランキング取得中')
        BillboadRank(Oriconday)
        update_progress_bar(window['progressbar'],window['progmsg'], 48,'DB登録中')
        asyncio.run(insertOriconWeekData())    
        asyncio.run(insertOriconDigitalData())
        asyncio.run(insertBillboardData())
        
        window.close()

        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'キャンセル':
                WriteLog(1,"得点取得画面を閉じる操作")
                break

    window.close()

  

def GetSelectWeekRank(SelectDay,Flag):

    WriteLog(1,"任意週データ取得を選択")

    # レイアウト（共通設定）
    layout = [
        [sg.Text('読み込み中...', size=(15, 1)), sg.ProgressBar(72, orientation='h', size=(20, 20), key='progressbar')],
        [sg.Button('読み込み中止'),sg.Text(key = 'progmsg')]
    ]


    window = sg.Window('任意週のランキング取得', layout,finalize=True,icon='FM-BACS.ico')

    def update_progress_bar(progress_bar,progmsg, value,msg):
        progress_bar.update_bar(value)
        progmsg.update(msg)
        window.refresh()

    while True:
        update_progress_bar(window['progressbar'],window['progmsg'],0,'データリセット中')
        clear()
        update_progress_bar(window['progressbar'],window['progmsg'], 8,'日付取得中')
        global OSW 
        OSW = OriconSelectWeek(SelectDay)
        if Flag == False:
            with open('Reload.txt', 'w',encoding='UTF-8') as f:
                SelectDay,Selecttime = str(SelectDay).split()
                f.write('3,,'+str(SelectDay)+',')
        update_progress_bar(window['progressbar'],window['progmsg'], 16,str(OSW)+'付けオリコン週間ランキング取得中')
        OriconWeekRank(OSW)
        update_progress_bar(window['progressbar'],window['progmsg'], 24,str(OSW)+'付けオリコンデジタルランキング取得中')
        OriconDigitalRank(OSW)
        update_progress_bar(window['progressbar'],window['progmsg'], 32,str(OSW - datetime.timedelta(days=5))+'付けビルボードランキング取得中')
        BillboadRank(OSW)
        update_progress_bar(window['progressbar'],window['progmsg'], 48,'DB登録中')
        asyncio.run(insertOriconWeekData())    
        asyncio.run(insertOriconDigitalData())
        asyncio.run(insertBillboardData())
        
        window.close()

        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'キャンセル':
                WriteLog(1,"得点取得画面を閉じる操作")
                break

    window.close()

def GetSelectWeekDate():
   return OSW


def OriconWeekUrl():
    Oriconday = OriconTodays()

    while True:
            # print(count)
            # print(Oriconday)
            load_url = "https://www.oricon.co.jp/rank/js/w/" + str(Oriconday) + "/"
            html = requests.get(load_url)
            soup = BeautifulSoup(html.text, 'lxml')
            main_content = soup.find(class_="content-rank-main")
            if main_content:
                break

            # 日付を一日減らす
            Oriconday = Oriconday - datetime.timedelta(days=1)

            # count = count + 1

    return load_url

def OriconDigitalUrl():
    Oriconday = OriconTodays()

    while True:
            # print(count)
            # print(Oriconday)
            load_url = "https://www.oricon.co.jp/rank/dis/w/" + str(Oriconday) + "/"
            html = requests.get(load_url)
            soup = BeautifulSoup(html.text, 'lxml')
            main_content = soup.find(class_="content-rank-main")
            if main_content:
                break

            # 日付を一日減らす
            Oriconday = Oriconday - datetime.timedelta(days=1)

            # count = count + 1

    return load_url

def BillboardUrl():
    Oriconday = OriconTodays()
        # URLの日付はOricondayと同じ

    while True:
        
        #URL(ここを変更すると読み込まなくなります)
        load_url = 'https://www.billboard-japan.com/charts/detail?a=hot100&year='+str(Oriconday.year)+'&month='+str(Oriconday.month)+'&day='+str(Oriconday.day)
        #URLを取得してくる
        response = requests.get(load_url)
        soup = BeautifulSoup(response.text, 'html.parser')
        songs = soup.find_all('p', class_='musuc_title') #曲名
        if songs:
            break
            # count = count + 1

        Oriconday = Oriconday - datetime.timedelta(days=1)

    return load_url

def ResetData():
    cursor.execute('UPDATE music_master SET Score = 0;') #Scoreの値を全消し
    cursor.execute('''DELETE FROM music_master WHERE Last_Number = '' OR Last_Number = 0;''') #Last_Numberの値が0もしくは空文字の場合消す
    print('DB関連処理終了')
    WriteLog(2,"DBのScoreを0に・ランキング外データを削除")
    conn.commit()

def GetLastNumber():
     # クエリの実行
    query = "SELECT MAX(Last_Number) FROM music_master;"
    cursor.execute(query)
    # 結果の取得
    max_last_number = cursor.fetchone()[0]
    last_number = int(max_last_number)

    return last_number

def WriteLog(label,logmessage):
    # コントロールログ（操作履歴を残す）
    # ラベル：ラベル0→起動、ラベル1→操作ログ（ボタン押したなど）、ラベル2→実行ログ、ラベル3→エラーログ
    log = open("MasterLog.log",'a') # ファイルオープン

    # 現在時刻を取得
    datetime_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

    if(label == 0):
        log.write("\n"+str(datetime_str)+" 起動：" +logmessage+"\n")

    elif(label == 1):
        log.write(str(datetime_str)+" 操作：" +logmessage + "\n")

    elif(label == 2):
        log.write(str(datetime_str)+" 実行：" +logmessage + "\n")

    elif(label == 4):
        log.write(str(datetime_str)+" エラー：" +logmessage+"\n")

    elif(label == 5):
        log.write(str(datetime_str)+" 終了：" +logmessage+"\n")

    log.close()

