from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
import openpyxl
import mojimoji
import sqlite3
import csv
import PySimpleGUI as sg
import GetData
import unicodedata


dbname = ('test2.db')
conn = sqlite3.connect(dbname, isolation_level=None)#データベースを作成、自動コミット機能ON
cursor = conn.cursor() #カーソルオブジェクトを作成

# 2022年1月16日分からのデータ取得
excel_file = 'Rank_BackUp/2ランキング.xlsx'
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

for low in range(2263, 4469,45):
    
    this_number = unicodedata.normalize("NFKC",sheet['B'+str(low+1)].value)
    this_number = this_number.replace('No.','')
        # データの処理と挿入
    for row in range(low+4, low+42, 2):
        title = unicodedata.normalize("NFKC", str(sheet['E' + str(row)].value))
        artist = unicodedata.normalize("NFKC", str(sheet['F' + str(row)].value))
        rank = sheet['B' + str(row)].value
        on_chart = sheet['D'+str(row)].value
        unique_id = GetData.generate_unique_id(title,artist)

        if "再" in str(sheet['C' + str(row)].value) or "圏外" in str(sheet['C' + str(row)].value) :
            last_number = int(this_number) - 1
        elif "初" in str(sheet['C' + str(row)].value):
            on_chart = 1
            last_number = this_number
        else:
            last_number = this_number

        # データベースに挿入
        cursor.execute('''INSERT INTO music_master
        (Title, Artist, Score, Last_Rank, Last_Number, On_chart,Unique_id)
        VALUES (?, ?, 0.0, ?, ?, ?, ?)
        ON CONFLICT(Unique_id) DO UPDATE SET Last_Number = ?,Last_Rank = ?,On_Chart = ?''',
                (title, artist, rank, last_number, on_chart,unique_id,last_number,rank,on_chart))

    # コミットとクローズ
    conn.commit()

    #Scoreの値を全消し
    cursor.execute("UPDATE music_master SET Score = 0;") 

    #Last_Numbweが空文字ものを全消し
    cursor.execute("DELETE FROM music_master WHERE Last_Number = '' OR Last_Number = 0;") 

    with open('楽曲データ.csv', 'w', newline='',encoding='UTF-8') as f:
        # CSVライターオブジェクトを作成
        csv_writer = csv.writer(f)

        # データを取得
        cursor.execute("SELECT * FROM music_master")
        rows = cursor.fetchall()

        # データをCSVに書き込む
        for row in rows:
            csv_writer.writerow(row)

    # sg.popup_ok('CSVファイルに書き込みました',no_titlebar=True)

import os
import pandas as pd
from datetime import datetime

def extract_date_from_filename(filename):
    try:
        # ファイル名から日付部分（YYYY-MM-DD）を抽出
        date_str = filename.split('ベストヒットランキング')[0]
        # 日付文字列をdatetimeオブジェクトに変換
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return None

def process_excel_files(directory):
    # 指定されたディレクトリ内のすべてのファイルを取得
    files = os.listdir(directory)
    
    # 指定された形式のExcelファイルのみをフィルタリング
    excel_files = [f for f in files if f.endswith('ベストヒットランキング.xlsx')]
    
    # ファイル名から日付を抽出し、タプル（ファイル名, 日付）のリストを作成
    dated_files = [(f, extract_date_from_filename(f)) for f in excel_files]
    
    # 有効な日付を持つファイルのみを保持
    valid_dated_files = [f for f in dated_files if f[1] is not None]
    
    # 日付順にソート
    sorted_files = sorted(valid_dated_files, key=lambda x: x[1])
    
    # 各ファイルを処理
    for filename, file_date in sorted_files:
        filepath = os.path.join(directory, filename)
        print(f"Processing file: {filename} with date: {file_date}")
        
        import RevisionRank2
        RevisionRank2.RevisionRank('Rank_BackUp/'+str(filename))

# 使用例
directory_path = 'Rank_BackUp'  # ここにExcelファイルが保存されているディレクトリのパスを指定
process_excel_files(directory_path)