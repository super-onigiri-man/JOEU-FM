from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
import sqlite3
from datetime import datetime
import PySimpleGUI as sg
import sys
def MajicalExcel(Oriconday):

    dbname = ('C:\\Users\\wiiue\\JOEU-FM\\test.db')
    conn = sqlite3.connect(dbname, isolation_level=None)#データベースを作成、自動コミット機能ON
    cursor = conn.cursor() #カーソルオブジェクトを作成

    # Excelファイルを読み込み
    workbook = load_workbook('ベストヒットランキング フォーマット.xlsx')
    sheet = workbook.active

    #Scoreの高い順に上位20件を取得
    top_20_query = '''
    SELECT * FROM music_master
    WHERE Score IS NOT NULL AND Score != ''
    ORDER BY Score DESC
    LIMIT 20
    '''

    top_20_results = cursor.execute(top_20_query).fetchall()

    # コミットして変更を保存
    conn.commit()

    date_string = str(Oriconday)
    date_obj = datetime.strptime(date_string, "%Y-%m-%d")
    Oriconday = int(date_obj.strftime("%Y%m%d"))

    year = Oriconday // 10000
    month = (Oriconday % 10000) // 100
    day = Oriconday % 100 

    # 最終回の確認
    # クエリの実行
    query = "SELECT MAX(Last_Number) FROM music_master;"
    cursor.execute(query)
    # 結果の取得
    max_last_number = cursor.fetchone()[0]
    # 今回のベストヒットランキング回数の値を設定
    this_rank_number = int(sg.popup_get_text('ベストヒットランキングのNoを入力してください\n現在、DBに登録されている最新Noは'+str(max_last_number)+'です', '曲名'))
    # print("ベストヒットランキングのNoを入力してください")
    # print("現在、DBに登録されている最新Noは"+str(max_last_number)+"です")
    # this_rank_number = int(input())
    rank = 1
    # 見出しのNo.と日付
    sheet.cell(row=3, column=2).value = "Ｎｏ."+str(+this_rank_number) #No.の書き込み
    sheet.merge_cells(start_row = 3 ,start_column=2 ,end_row =3 , end_column = 4) #セル結合

    sheet.cell(row=3, column=6).value = str(year)+"年"+str(month)+"月"+str(day)+"日"#日付の書き込み

    # ランキングセルにデータを書き込み
    for idx in range(1, len(top_20_results)*2, 2):
        result = top_20_results[idx // 2]  # idxを2で割った商を使って結果を取得
        title, artist, score, last_rank, last_number, on_chart= result

        # ランキング初登場曲
        # Last_Rankが空文字もしくはnullの場合、文字色を赤に変更
        if last_rank is None or last_rank == "" or last_rank == 0:
            sheet.cell(row=idx+5, column=2).value = rank
            sheet.cell(row=idx+5, column=2).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="FF0000") #ThisWeek
            sheet.alignment = Alignment(horizontal = 'center',vertical='center')
            sheet.merge_cells(start_row = idx +5 ,start_column=2 ,end_row = idx +6 , end_column = 2)
            sheet.cell(row=idx+5, column=3).value = "初"
            sheet.cell(row=idx+5, column=3).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="FF0000") #LastWeek
            sheet.merge_cells(start_row = idx +5 ,start_column=3 ,end_row = idx +6 , end_column = 3)
            sheet.cell(row=idx+5, column=4).value = "1"
            sheet.cell(row=idx+5, column=4).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="FF0000") #Onchart
            sheet.merge_cells(start_row = idx +5 ,start_column=4 ,end_row = idx +6 , end_column = 4)
            sheet.cell(row=idx+5, column=5).value = title
            sheet.cell(row=idx+5, column=6).value = artist

            rank += 1

        # ランキング再登場曲
        # Last_NumberがThis_Numberより2以上差がある場合、セルに「再」を入力
        elif (this_rank_number - last_number) >= 2 :
            sheet.cell(row=idx+5, column=2).value = rank#ThisWeek
            sheet.cell(row=idx+5, column=2).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="000000")
            sheet.merge_cells(start_row = idx +5 ,start_column=2 ,end_row = idx +6 , end_column = 2)
            sheet.cell(row=idx+5, column=3).value = "再"#LastWeek
            sheet.cell(row=idx+5, column=3).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="000000")
            sheet.merge_cells(start_row = idx +5 ,start_column=3 ,end_row = idx +6 , end_column = 3)
            sheet.cell(row=idx+5, column=4).value = on_chart+1
            sheet.cell(row=idx+5, column=4).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="000000")
            sheet.merge_cells(start_row = idx +5 ,start_column=4 ,end_row = idx +6 , end_column = 4)
            sheet.cell(row=idx+5, column=5).value = title
            sheet.cell(row=idx+5, column=6).value = artist

            rank+=1

        else:
        #  先週も登場した曲
        # セルにデータを書き込み
            sheet.cell(row=idx+5, column=5).value = title
            sheet.cell(row=idx+5, column=6).value = artist
            sheet.cell(row=idx+5, column=2).value = rank
            sheet.merge_cells(start_row = idx +5 ,start_column=2 ,end_row = idx +6 , end_column = 2)
            sheet.cell(row=idx+5, column=2).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="000000")
            sheet.cell(row=idx+5, column=3).value = last_rank
            sheet.cell(row=idx+5, column=3).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="000000")
            sheet.merge_cells(start_row = idx +5 ,start_column=3 ,end_row = idx +6 , end_column = 3)
            # sheet.cell(row=idx, column=7).value = last_number
            sheet.cell(row=idx+5, column=4).value = on_chart+1
            sheet.cell(row=idx+5, column=4).font = Font(name ="BIZ UDPゴシック",size = 16,bold=True,color="000000")
            sheet.merge_cells(start_row = idx +5 ,start_column=4 ,end_row = idx +6 , end_column = 4)
            rank+=1

    # Excelファイルを保存
    workbook.save(str(year)+'-'+str(+month)+'-'+str(day)+'ベストヒットランキング.xlsx')

    result = sg.popup_ok('正常に処理されました')

    return 0
