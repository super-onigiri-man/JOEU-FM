from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
import openpyxl
import unicodedata
import sqlite3
import csv
import PySimpleGUI as sg
import GetData
import shutil
import datetime


dbname = ('test.db')
conn = sqlite3.connect(dbname, isolation_level=None)#データベースを作成、自動コミット機能ON
cursor = conn.cursor() #カーソルオブジェクトを作成

def ExcelCheck(RevisionPath):

    GetData.WriteLog(3,"ランキング修正：Excelチェックを実行")

    try:
        # Excelファイルの読み込み
        excel_file = RevisionPath
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # クエリの実行
        query = "SELECT MAX(Last_Number) FROM music_master;"
        cursor.execute(query)
        # 結果の取得
        max_last_number = cursor.fetchone()[0]

        if sheet['B3'].value is None or sheet['F3'].value is None:
            sg.popup_error('回数もしくは日付が入力されていません',no_titlebar=True)

            return False

        else:
            this_number = unicodedata.normalize("NFKC",sheet['B3'].value)
            this_number = this_number.replace('No.','')

            if int(max_last_number) > int(this_number):
                sg.popup_error('ランキングは最新回のみ登録できます',no_titlebar=True)
                return False
            
            else:
                return True
            
    except Exception as e:
            import traceback
            with open('error.log', 'a') as f:
                traceback.print_exc(file=f)
            GetData.WriteLog(4,"ランキング修正：Excelチェックに失敗")
            sg.popup_error("ランキングが取得できませんでした\nベストヒットランキングを登録してください", title="エラー",no_titlebar=True)



def RevisionRank(RevisionPath):
    GetData.WriteLog(3,"ランキング修正：ランキング修正を実行")
    # Excelファイルの読み込み
    excel_file = RevisionPath
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    this_number = unicodedata.normalize("NFKC",sheet['B3'].value)
    this_number = this_number.replace('No.','')

        # データの処理と挿入
    for row in range(6, 45, 2):
        title = unicodedata.normalize("NFKC",str(sheet['E' + str(row)].value))
        artist = unicodedata.normalize("NFKC",str(sheet['F' + str(row)].value))
        rank = sheet['B' + str(row)].value
        on_chart = sheet['D'+str(row)].value
        unique_id = GetData.generate_unique_id(title,artist)
        
        if "再" in str(sheet['C' + str(row)].value) or "圏外" in str(sheet['C' + str(row)].value) :
            last_number = int(this_number) - 1
        elif "初" in str(sheet['C' + str(row)].value):
            on_chart = 1
            last_number = this_number
        else:
            last_number = this_number

        # データベースに挿入
        cursor.execute('''INSERT INTO music_master
        (Title, Artist, Score, Last_Rank, Last_Number, On_chart,Unique_id)
        VALUES (?, ?, 0.0, ?, ?, ?, ?)
        ON CONFLICT(Unique_id) DO UPDATE SET Last_Number = ?,Last_Rank = ?,On_Chart = ?''',
                (title, artist, rank, last_number, on_chart,unique_id,last_number,rank,on_chart))

    # コミットとクローズ
    conn.commit()

    #Scoreの値を全消し
    cursor.execute("UPDATE music_master SET Score = 0;") 

    #Last_Numbweが空文字ものを全消し
    cursor.execute("DELETE FROM music_master WHERE Last_Number = '' OR Last_Number = 0;") 

    with open('楽曲データ.csv', 'w', newline='',encoding='UTF-8') as f:
        # CSVライターオブジェクトを作成
        csv_writer = csv.writer(f)

        # データを取得
        cursor.execute("SELECT * FROM music_master")
        rows = cursor.fetchall()

        # データをCSVに書き込む
        for row in rows:
            csv_writer.writerow(row)

    GetData.WriteLog(3,"ランキング修正：Excelチェックに成功")
    sg.popup_ok('CSVファイルに書き込みました',no_titlebar=True)

def CopyFile(RevisionPath):

    # Excelファイルの読み込み
    excel_file = RevisionPath
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    this_date = unicodedata.normalize("NFKC",sheet['F3'].value)

    # 日付文字列をdatetimeオブジェクトに変換
    date_object = datetime.datetime.strptime(this_date, '%Y年%m月%d日')

    # datetimeオブジェクトを指定の形式の文字列に変換
    formatted_date = date_object.strftime('%Y-%m-%d')

    
    shutil.copy(RevisionPath, 'Rank_BackUp/'+str(formatted_date)+'ベストヒットランキング.xlsx')
    GetData.WriteLog(3,"ランキング修正：Rank_BackUp/"+str(formatted_date)+"ベストヒットランキング.xlsxをコピーを実行")